import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side

import matplotlib.pyplot as plt
import japanize_matplotlib

import json
import csv
import os

def make_output_xlsx(file_name):
    def get_average(last_row, column):
        total = 0
        count = 0
        
        for row in range(25, last_row+1):
            cell_value = sheet.cell(row=row, column=column).value
            total += cell_value
            count += 1
        
        return total / count if count > 0 else None

    with open("input.json", 'r') as f:
        data = json.load(f)

    b_column_data = {
        "初期タンク圧": data["oxidizer"]["Initial tank pressure [Nm^-2]"] / 1e6,
        "最終タンク圧": data["oxidizer"]["Final tank pressure [Nm^-2]"] / 1e6,
        "酸化剤充填量": data["oxidizer"]["Oxidizer filling volume [m^3]"] * 1e6,
        "酸化剤密度": data["oxidizer"]["Oxidizer density [kgm^-3]"],
        "燃料密度": data["fuel"]["Fuel density [kgm^-3]"],
        "燃料軸長": data["fuel"]["Fuel length [m]"] * 1e3,
        "初期ポート径": data["fuel"]["Initial port diameter [m]"] * 1e3,
        "燃料外径": data["fuel"]["Fuel outer diameter [m]"] * 1e3,
        "ポート数": data["fuel"]["Fuel port number [-]"],
        "オリフィス径": data["Oxidizer flow characteristics"]["Orifice diameter [m]"] * 1e3,
        "流量係数": data["Oxidizer flow characteristics"]["Flow coefficient [-]"],
        "酸化剤流束係数": data["Combustion characteristics"]["Oxidizer mass flux coefficient [m^3kg^-1]"],
        "酸化剤流束指数": data["Combustion characteristics"]["Oxidizer mass flux exponent [-]"],
        "c*効率": data["Combustion characteristics"]["C-star efficiency [-]"],
        "初期スロート径": data["Nozzle characteristics"]["Initial nozzle throat diameter [m]"] * 1e3,
        "ノズル出口径": data["Nozzle characteristics"]["Nozzle exit diameter [m]"] * 1e3,
        "ノズル開口半頂角": data["Nozzle characteristics"]["Nozzle exit half angle [deg]"],
        "エロージョン速度": data["Nozzle characteristics"]["Nozzle erosion speed [ms^-1]"],
        "背圧": data["Environment"]["Back pressure [Nm^-2]"] / 1e6,
    }

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    with open("output.csv", 'r') as f:
        csv_reader = csv.reader(f)
        for row_num, row_data in enumerate(csv_reader, start=24):
            for col_num, cell_value in enumerate(row_data, start=1):
                try:
                    cell_value = float(cell_value) if 'e' in cell_value or '.' in cell_value or cell_value.isdigit() else cell_value
                    if col_num in [2, 3, 4]:
                        cell_value = cell_value / 1e6
                    if col_num in [11, 12, 14]:
                        cell_value = cell_value * 1e3
                except ValueError:
                    pass
                sheet.cell(row=row_num, column=col_num).value = cell_value

    index_row = ["時刻", "タンク圧", "燃焼室圧", "ノズル出口圧", "酸化剤消費量", "燃料消費量", "酸化剤質量流量", "燃料質量流量", "O/F", "酸化剤質量流束", "ポート径", "燃料後退速度", "特性排気速度", "スロート径", "比熱比", "推力係数", "開口比", "推力", "トータルインパルス"]
    unit_row = ["s", "MPa", "MPa", "MPa", "kg", "kg", "kg/s", "kg/s", "-", "kg/m3s", "mm", "mm/s", "m/s", "mm", "-", "-", "-", "N", "Ns"]

    for col_num in range(len(index_row)):
        sheet.cell(row=23, column=col_num+1).value = index_row[col_num]
        sheet.cell(row=24, column=col_num+1).value = unit_row[col_num]
        col_letter = get_column_letter(col_num+1)
        sheet.column_dimensions[col_letter].width = 15

    last_row = sheet.max_row
    columns = {
        "A": ["入力値", "初期タンク圧", "最終タンク圧", "酸化剤充填量", "酸化剤密度", "燃料密度", "燃料軸長", "初期ポート径", "燃料外径", "ポート数", "オリフィス径", "流量係数", "酸化剤流束係数", "酸化剤流束指数", "c*効率", "初期スロート径", "ノズル出口径", "ノズル開口半頂角", "エロージョン速度", "背圧"],
        "D": [None, "MPa", "MPa", "cc", "kg/m3", "kg/m3", "mm", "mm", "mm", "-", "mm", "-", "m3/kg", "-", "-", "mm", "mm", "deg", "mm/s", "MPa"],
        "F": ["出力値", "燃焼時間", "トータルインパルス", "平均推力", "酸化剤消費量", "燃料消費量", "比推力", "最終燃料ポート径", "スライバ率", "平均タンク圧力", "平均燃焼室圧", "平均ノズル出口圧", "平均酸化剤流量", "平均燃料流量", "平均燃料後退速度", "平均O/F", "平均特性排気速度", "平均推力係数", "最終スロート径", "初期燃焼室特性長", "最終燃焼室特性長"],
        "H": [None, 
              sheet.cell(row=last_row, column=1).value, 
              sheet.cell(row=last_row, column=19).value, 
              get_average(last_row, 18), 
              sheet.cell(row=last_row, column=5).value, 
              sheet.cell(row=last_row, column=6).value, 
              sheet.cell(row=last_row, column=19).value/((sheet.cell(row=last_row, column=5).value+sheet.cell(row=last_row, column=6).value)*9.8),
              sheet.cell(row=last_row, column=11).value,
              (b_column_data["燃料外径"]**2-b_column_data["ポート数"]*sheet.cell(row=last_row, column=11).value**2)/(b_column_data["燃料外径"]**2-b_column_data["ポート数"]*b_column_data["初期ポート径"]**2)*100,
              get_average(last_row, 2),
              get_average(last_row, 3),
              get_average(last_row, 4),
              get_average(last_row, 7),
              get_average(last_row, 8),
              get_average(last_row, 12),
              get_average(last_row, 9),
              get_average(last_row, 13),
              get_average(last_row, 16),
              sheet.cell(row=last_row, column=14).value,
              b_column_data["燃料軸長"]*b_column_data["ポート数"]*(b_column_data["初期ポート径"]/sheet.cell(row=last_row, column=14).value)**2/1e3,
              b_column_data["燃料軸長"]*b_column_data["ポート数"]*(sheet.cell(row=last_row, column=11).value/sheet.cell(row=last_row, column=14).value)**2/1e3
              ],
        "I": [None, "s", "Ns", "N", "kg", "kg", "s", "mm", "%", "MPa", "MPa", "MPa", "kg/s", "kg/s", "mm/s", "-", "m/s", "-", "mm", "m", "m"],
    }

    for col, values in columns.items():
        for row_num, value in enumerate(values, start=1):
            if isinstance(value, (int, float)):
                sheet[f"{col}{row_num}"].value = round(value, 3)
            else:
                sheet[f"{col}{row_num}"].value = value

    for row_num, key in enumerate(columns["A"][1:], start=2):
        if key in b_column_data:
            sheet.cell(row=row_num, column=3).value = b_column_data[key]

    for row in range(2, 22):
        if row != 22:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = sheet.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal="left", vertical="center")

        sheet.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        cell = sheet.cell(row=row, column=1)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    put_pressure_fig(sheet, last_row, "B26")
    put_of_fig(sheet, last_row, "H26")

    side1 = Side(style='thin', color='000000')
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)

    for row in sheet['A2:D20']:
        for cell in row:
            cell.border = border_aro
    
    for row in sheet['F2:I21']:
        for cell in row:
            cell.border = border_aro

    if file_name == '':
        file_name = "result.xlsx"
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    workbook.save(file_name)

    if os.path.exists("chart_pressure.png"):
        os.remove("chart_pressure.png")
    if os.path.exists("chart_of.png"):
        os.remove("chart_of.png")

def put_pressure_fig(sheet, last_row, print_cell):
    x_values = [sheet.cell(row=i, column=1).value for i in range(25, last_row + 1)]
    y1_values = [sheet.cell(row=i, column=2).value for i in range(25, last_row + 1)]
    y2_values = [sheet.cell(row=i, column=3).value for i in range(25, last_row + 1)]
    y3_values = [sheet.cell(row=i, column=4).value for i in range(25, last_row + 1)]
    y4_values = [sheet.cell(row=i, column=18).value for i in range(25, last_row + 1)]

    _, ax1 = plt.subplots(figsize=(6, 4))
    line1, = ax1.plot(x_values, y1_values, label="タンク圧", color="blue")
    line2, = ax1.plot(x_values, y2_values, label="燃焼室圧", color="green")
    line3, = ax1.plot(x_values, y3_values, label="ノズル出口圧", color="red")
    ax1.set_xlim([0, max(x_values)])
    ax1.set_ylim([0, max(y1_values) + 1])
    ax1.set_xlabel("時刻 [s]")
    ax1.set_ylabel("圧力 [MPa]")
    ax1.grid(True)

    ax2 = ax1.twinx()
    line4, = ax2.plot(x_values, y4_values, label="推力", color="orange")
    ax2.set_ylim([0, max(y4_values) + 100])
    ax2.set_ylabel("推力 [N]")

    lines = [line1, line2, line3, line4]
    labels = [line.get_label() for line in lines]
    ax1.legend(lines, labels, loc="upper right")

    plt.savefig("chart_pressure.png")
    plt.close()

    img = Image("chart_pressure.png")
    sheet.add_image(img, print_cell)

def put_of_fig(sheet, last_row, print_cell):
    x_values = [sheet.cell(row=i, column=1).value for i in range(25, last_row + 1)]
    y1_values = [sheet.cell(row=i, column=13).value for i in range(25, last_row + 1)]
    y2_values = [sheet.cell(row=i, column=9).value for i in range(25, last_row + 1)]

    _, ax1 = plt.subplots(figsize=(6, 4))
    line1, = ax1.plot(x_values, y1_values, label="特性排気速度", color="blue")
    ax1.set_xlim([0, max(x_values)])
    ax1.set_ylim([0, max(y1_values) * 1.1])
    ax1.set_xlabel("時刻 [s]")
    ax1.set_ylabel("特性排気速度 [m/s]")
    ax1.grid(True)

    ax2 = ax1.twinx()
    line2, = ax2.plot(x_values, y2_values, label="O/F", color="red")
    ax2.set_ylim([0, max(y2_values) + 2])
    ax2.set_ylabel("O/F [-]")

    lines = [line1, line2]
    labels = [line.get_label() for line in lines]
    ax1.legend(lines, labels, loc="lower right")

    plt.savefig("chart_of.png")
    plt.close()

    img = Image("chart_of.png")
    sheet.add_image(img, print_cell)

if __name__ == "__main__":
    make_output_xlsx("test.xlsx")
